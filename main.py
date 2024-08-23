import dash
from dash import html, dcc
from dash.dependencies import Input, Output, State
import pandas as pd
import os
import glob
import plotly.graph_objs as go
import numpy as np
import bell_contour
import io
import rocket_calculator
import shutil
import xlsxwriter
from openpyxl import load_workbook

# Crear la aplicación Dash
external_stylesheets = ['/assets/style.css']
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
app.config.suppress_callback_exceptions = True

# Función para listar archivos en una carpeta específica
def list_files():
    file_names = [os.path.basename(file).split('.')[0] for file in glob.glob(os.path.join(os.getcwd()+'/configurations',f'*.xlsx'))]
    return file_names

# Layout de la página 2
settings_layout = html.Div(
    className="content",
    children=[
        dcc.Store(id='page-2-store'),
        html.Div(className="navbar",
                    children=[
                                html.Div(className="navbar-title",  children=["Rocket Lab", html.Img(src='/assets/rocket_icon.png', className='navbar-icon')]),
                                html.Div(className="navbar-links", children=[
                                                                            dcc.Link('Simulations', href='/simulations', className='navbar-link'),
                                                                            dcc.Link('Settings', href='/settings', className='navbar-link'),
                                                                            ])
                            ]
                ),
        html.Div(className="title-container", children=[
            html.Label('Settings', className='title-label'),
        ]),
        html.Div(className="input-block-container", children=[
            html.Div(className="input-block", children=[
                html.Div(className='input-label-div', children=[
                    html.Label('Propellant parameters', className='input-label'),
                ]),
                html.Div(id='Tc-div', className='div-input', children=[
                    html.Label('Combustion temperature: ', className='name-input-box'),
                    dcc.Input(id='Tc-input', type='number', min=0, className='input-box'),
                    html.Label('K', className='units-input-box')
                ]),
                html.Div(id='Masa_molar-div', className='div-input', children=[
                    html.Label('Molar mass: ', className='name-input-box'),
                    dcc.Input(id='Masa_molar-input', type='number', min=0, className='input-box'),
                    html.Label('g/mol', className='units-input-box')
                ]),
                html.Div(id='gamma-div', className='div-input', children=[
                    html.Label('Gamma:', className='name-input-box'),
                    dcc.Input(id='gamma-input', type='number', min=0, className='input-box'),
                    html.Label('-', className='units-input-box')
                ]),
                html.Div(id='Rend-div', className='div-input', children=[
                    html.Label('Solid to gas efficiency: ', className='name-input-box'),
                    dcc.Input(id='Rend-input', type='number', min=0, className='input-box'),
                    html.Label('%', className='units-input-box')
                ]),
                html.Div(id='rho_pr-div', className='div-input', children=[
                    html.Label('Propellant density:', className='name-input-box'),
                    dcc.Input(id='rho_pr-input', type='number', min=0, className='input-box'),
                    html.Label('kg/m³', className='units-input-box')
                ]),
                html.Div(id='a-div', className='div-input', children=[
                    html.Label('a: ', className='name-input-box'),
                    dcc.Input(id='a-input', type='number', min=0, className='input-box'),
                    html.Label('', className='units-input-box')
                ]),
                html.Div(id='n-div', className='div-input', children=[
                    html.Label('n: ', className='name-input-box'),
                    dcc.Input(id='n-input', type='number', min=0, className='input-box'),
                    html.Label('-', className='units-input-box')
                ]),
            ]),
            html.Div(className='input-block', children=[
                html.Div(className="input-block", children=[
                    html.Div(className='input-label-div', children=[
                        html.Label('Rocket Parameters', className='input-label'),
                    ]),
                    html.Div(id='Mcp-div', className='div-input', children=[
                        html.Label('Payload mass:', className='name-input-box'),
                        dcc.Input(id='Mcp-input', type='number', min=0, className='input-box'),
                        html.Label('kg', className='units-input-box')
                    ]),
                    html.Div(id='Mc-div', className='div-input', children=[
                        html.Label('Engine mass:', className='name-input-box'),
                        dcc.Input(id='Mc-input', type='number', min=0, className='input-box'),
                        html.Label('kg', className='units-input-box')
                    ]),
                    html.Div(id='Rcg_cp-div', className='div-input', children=[
                        html.Label('Mass center: ', className='name-input-box'),
                        dcc.Input(id='Rcg_cp-input', type='number', min=0, className='input-box'),
                        html.Label('mm', className='units-input-box')
                    ]),
                    html.Div(id='Cd-div', className='div-input', children=[
                        html.Label('Drag coefficient:', className='name-input-box'),
                        dcc.Input(id='Cd-input', type='number', min=0, className='input-box'),
                        html.Label('-', className='units-input-box')
                    ]),
                ]),
                html.Div(className="input-block", children=[
                    html.Div(className='input-label-div', children=[
                        html.Label('Constants', className='input-label'),
                    ]),
                    html.Div(id='Re-div', className='div-input', children=[
                        html.Label('Earth radius:', className='name-input-box'),
                        dcc.Input(id='Re-input', type='number', min=0, className='input-box'),
                        html.Label('km', className='units-input-box')
                    ]),
                    html.Div(id='g0-div', className='div-input', children=[
                        html.Label('Gravity intensity on surface:', className='name-input-box'),
                        dcc.Input(id='g0-input', type='number', min=0, className='input-box'),
                        html.Label('m/s²', className='units-input-box')
                    ]),
                    html.Div(id='Ra-div', className='div-input', children=[
                        html.Label('Gas constant of air:', className='name-input-box'),
                        dcc.Input(id='Ra-input', type='number', min=0, className='input-box'),
                        html.Label('J/(kg·K)', className='units-input-box')
                    ]),
                ]),
            ]),
        ]),
        html.Div(className="input-block", children=[
            html.Div(className='input-label-div', children=[
                html.Label('Simulations parameters', className='input-label'),
            ]),
            html.Div(className='input-block2', children=[
                html.Div(id='t0-div', className='div-input', children=[
                    html.Label('Initial time simulation:', className='name-input-box'),
                    dcc.Input(id='t0-input', type='number', min=0, className='input-box'),
                    html.Label('s', className='units-input-box')
                ]),
                html.Div(id='tf-div', className='div-input', children=[
                    html.Label('Final time simulation:', className='name-input-box'),
                    dcc.Input(id='tf-input', type='number', min=0, className='input-box'),
                    html.Label('s', className='units-input-box')
                ]),
                html.Div(id='N-div', className='div-input', children=[
                    html.Label('Number of steps:', className='name-input-box'),
                    dcc.Input(id='N-input', type='number', min=0, step=1, className='input-box'),
                    html.Label('-', className='units-input-box')
                ]),
            ])
        ]),
        html.Div(className='configuration-name-container', children=[
            html.Div(className='configuration-name-div', children=[
                html.Label('Configuration name:', className='config-label'),
                dcc.Input(id='name-input', type='text', placeholder='write name', className='config-input-box'),
            ])
        ]),
        html.Div(className='buttons-container', children=[
            html.Div(className='button-div', children=[
                html.Button('Load', id='open-modal-button', className='button-modern'),
            ]),
            html.Div(className='button-div', children=[
                html.Button('Save', id ='save-button', className='button-modern'),
            ])
        ]),
        html.Div(id='error-message', className='error-message-box'),
        html.Div(id='success-message', className='success-message-box'),
        html.Div(id='modal', className='modal', style={'display': 'none'}, children=[
            html.Div(className='modal-content', children=[
                html.H2('Select an Excel file'),
                dcc.Dropdown(id='file-dropdown', options=[{'label': f, 'value': f} for f in list_files()], placeholder="Select a file"),
                html.Div(className='modal-buttons', children=[
                    html.Button('Accept', id='accept-button', className='button-modern'),
                    html.Button('Cancel', id='cancel-button', className='button-modern')
                ])
            ])
        ]),
    ]
)

# Layout de la página 1 (ejemplo)
simulations_layout = html.Div(className="content",
                      children=[
                          dcc.Store(id='page-1-store'),
                          dcc.Store(id='page-2-store'),
                          html.Div(className="navbar",
                                   children=[
                                       html.Div(className="navbar-title", children=["Rocket Lab", html.Img(src='/assets/rocket_icon.png', className='navbar-icon')]),
                                       html.Div(className="navbar-links", children=[
                                           dcc.Link('Simulations', href='/simulations', className='navbar-link'),
                                           dcc.Link('Settings', href='/settings', className='navbar-link'),
                                       ])
                                   ]
                                   ),
                          html.Div(className="title-container",
                                   children=[
                                       html.Label('Simulations', className='title-label'),
                                   ]
                                   ),
                          html.Div(className='input-block-container',
                                   children=[
                                       html.Div(className="input-block",
                                                children=[
                                                    html.Div(className='configuration-name-container-simulations',
                                                             children=[
                                                                 html.Div(className='configuration-name-div',
                                                                          children=[
                                                                              html.Label('Settings configuration selected:',
                                                                                         className='config-label'),
                                                                              dcc.Input(id='name-input', type='text', readOnly=True,
                                                                                        className='config-input-box', style={'padding': '30px'}),
                                                                          ]
                                                                          )
                                                             ]),
                                                    html.Div(className='input-label-div',
                                                             children=[
                                                                 html.Label('Nozzle', className='input-label'),
                                                             ]
                                                             ),
                                                    html.Div(id='rg-div',
                                                             className='div-input', children=[
                                                         html.Label('Throat radius ', className='name-input-box'),
                                                         dcc.Input(id='rg-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('mm', className='units-input-box')
                                                     ]
                                                             ),
                                                    html.Div(id='exit-radius-div',
                                                             className='div-input', children=[
                                                         html.Label('Exit radius: ', className='name-input-box'),
                                                         dcc.Input(id='rs-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('mm', className='units-input-box')
                                                     ]
                                                             ),

                                                    html.Div(className='input-label-div',
                                                             children=[
                                                                 html.Label('Solid Fuel', className='input-label'),
                                                             ]
                                                             ),
                                                    html.Div(id='R-div',
                                                             className='div-input', children=[
                                                         html.Label('Final fuel radius:', className='name-input-box'),
                                                         dcc.Input(id='R-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('mm', className='units-input-box')
                                                     ]
                                                             ),
                                                    html.Div(id='R0-div',
                                                             className='div-input', children=[
                                                         html.Label('Initial fuel radius:', className='name-input-box'),
                                                         dcc.Input(id='R0-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('mm', className='units-input-box')
                                                     ]
                                                             ),

                                                    html.Div(id='L-div',
                                                             className='div-input', children=[
                                                         html.Label('Fuel length:', className='name-input-box'),
                                                         dcc.Input(id='L-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('mm', className='units-input-box')
                                                     ]
                                                             ),
                                                    
                                                    html.Div(className='input-label-div',
                                                             children=[
                                                                 html.Label('Data representation', className='input-label'),
                                                             ]
                                                             ),
                                                    html.Div(id='L-div',
                                                             className='div-input', children=[
                                                         html.Label('Points to save:', className='name-input-box'),
                                                         dcc.Input(id='N_sc-input', type='number', min=0,
                                                                   className='input-box'),
                                                         html.Label('', className='units-input-box')
                                                     ]
                                                             ),
                                                    html.Div(className='dropdown-nozzle-container',
                                                             children=[
                                                                 dcc.Dropdown(className='dropdown-nozzle',
                                                                              id='nozzle-type-dropdown',
                                                                              options=[
                                                                                  {'label': 'Bell nozzle', 'value': 'Bell nozzle'}],
                                                                              placeholder="Select a nozzle type",
                                                                              ),
                                                             ]
                                                             ),
                                                    
                                                    html.Div(className='buttons-container', 
                                                            children=[html.Div(className='button-div',
                                                                                children=[
                                                                                         html.Button('Calculate', id='calculate-button', className='button-modern'),
                                                                                        ]
                                                                                ),
                                                    html.Div(className='button-div',
                                                            children=[
                                                                        html.Button('Save', id ='open-modal2-button', className='button-modern', style={'display':'none'}),
                                                                    ]),
                                                    
                                                    html.Div(className='button-div',
                                                            children=[
                                                                        html.Button('Download nozzle contour', id ='download-nozzle-button', className='button-modern'),
                                                                        dcc.Download(id='download-nozzle-contour'),
                                                                    ])
                                                            ]),
                                                    
                                                    html.Div(id='success-saving-message', style={'color': 'green', 'margin-top': '10px', 'text-align': 'center'}),
                                                    

                                                ]
                                                ),
                                       html.Div(id='modal2',
                                                className='modal',
                                                style={'display': 'none'},
                                                children=[
                                                    html.Div(className='modal-content',
                                                            children=[
                                                                html.H2('Filename:'),
                                                                dcc.Input(id='save_filename-input', type='text', className='input-box'),
                                                                html.Div(id='name-exists-warning', style={'color': 'red', 'margin-top': '10px'}),
                                                                html.Div(className='modal-buttons',
                                                                        children=[
                                                                            html.Button('Accept', id='accept-button-modal2', className='button-modern'),
                                                                            html.Button('Cancel', id='cancel-button-modal2', className='button-modern')
                                                                        ]
                                                                )
                                                            ]
                                                    )
                                                ]
                                        ),
                                       html.Div(className="contour-graph-block",
                                                children=[
                                                    html.Div(className='checklist-container',
                                                    children=[
                                                    html.Div(className='checklist-div',
                                                            children=[
                                                                        dcc.Checklist( id='3d-toggle',
                                                                                    options=[{'label': '3D view', 'value': '3d'}],
                                                                                    value=[]
                                                                        )
                                                    ]
                                                    ),
                                                    html.Div(className='checklist-div',
                                                             children=[
                                                                        dcc.Checklist( id='show-axis-toggle',
                                                                                    options=[{'label': 'Show axis', 'value': 'axis'}],
                                                                                    value=[]
                                                                        
                                                                                ),
                                                             ]
                                                    ),
                                                    ]
                                                    ),
                                                    dcc.Graph(id='contour-graph',
                                                              config={'displayModeBar': False},
                                                               style={'height': '50vh'})
                                                ]

                                                )
                                   ]
                                   ),
                        html.Div(id='graphs-container',
                                 style={'display':'none'},
                                children=[
                                    html.Div(className='input-block-container',
                                            children=[
                                                html.Div(className='input-block', 
                                                        children=[
                                                            dcc.Graph(id='graph1', config={'displayModeBar': False}, style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph2', style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph3', style={'height': '40vh'}),
                                                        ])
                                            ]),
                                    html.Div(className='input-block-container',
                                            children=[
                                                html.Div(className='input-block', 
                                                        children=[
                                                            dcc.Graph(id='graph4', config={'displayModeBar': False}, style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph5', style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph6', style={'height': '40vh'}),
                                                        ])
                                            ]),
                                    html.Div(className='input-block-container',
                                            children=[
                                                html.Div(className='input-block', 
                                                        children=[
                                                            dcc.Graph(id='graph7', config={'displayModeBar': False}, style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph8', style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph9', style={'height': '40vh'}),
                                                        ])
                                            ]),
                                    
                                    html.Div(className='input-block-container',
                                            children=[
                                                html.Div(className='input-block', 
                                                        children=[
                                                            dcc.Graph(id='graph10', config={'displayModeBar': False}, style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph11', style={'height': '40vh'}),
                                                        ]),
                                                html.Div(className='input-block',
                                                        children=[
                                                            dcc.Graph(id='graph12', style={'height': '40vh'}),
                                                        ])
                                            ])
                                ])
                      ]
)

# Layout principal
app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    html.Div(id='page-content')
])

# Callback para enrutar a diferentes páginas
@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/settings':
        return settings_layout
    elif pathname == '/simulattions':
        return simulations_layout
    else:
        return simulations_layout

# Registrar los callbacks para la página 2
@app.callback(
    [Output(f'{id}-div', 'style') for id in [
        'Tc', 'Masa_molar', 'gamma', 'rho_pr', 'Rend', 'a', 'n',
        'Mcp', 'Rcg_cp', 'Cd', 'Re', 'g0', 'Ra', 'Mc', 't0', 'tf', 'N']] +
    [Output('error-message', 'children'), Output('success-message', 'children'), Output('modal', 'style'), Output('name-input', 'value'), Output('page-2-store', 'data')] +
    [Output(f'{id}-input', 'value') for id in [
        'Tc', 'Masa_molar', 'gamma', 'rho_pr', 'Rend', 'a', 'n',
        'Mcp', 'Rcg_cp', 'Cd', 'Re', 'g0', 'Ra', 'Mc', 't0', 'tf', 'N']] +
    [Output('file-dropdown', 'options')],
    [Input('save-button', 'n_clicks'), Input('open-modal-button', 'n_clicks'),
     Input('cancel-button', 'n_clicks'), Input('accept-button', 'n_clicks'), Input('url', 'pathname')],
    [State('modal', 'style'), State('page-2-store', 'data')] +
    [State(f'{id}-input', 'value') for id in [
        'Tc', 'Masa_molar', 'gamma', 'rho_pr', 'Rend', 'a', 'n',
        'Mcp', 'Rcg_cp', 'Cd', 'Re', 'g0', 'Ra', 'Mc', 't0', 'tf', 'N']] + [State('name-input', 'value'), State('file-dropdown', 'value')]
)
def handle_callbacks(save_clicks, open_clicks, cancel_clicks, accept_clicks, pathname, modal_style, page_2_store_data, *values):
    ctx = dash.callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]

    ids = [
        'Tc', 'Masa_molar', 'gamma', 'rho_pr', 'Rend', 'a', 'n',
        'Mcp', 'Rcg_cp', 'Cd', 'Re', 'g0', 'Ra', 'Mc', 't0', 'tf', 'N'
    ]
    styles = [{}] * len(ids)
    inputs_values = list(values[:-2])
    name_value = values[-2]
    file_dropdown_value = values[-1]
    success_message = ''
    error_message = ''
    
    if triggered_id == '':
        try:
            selected_file_name = page_2_store_data['selected_file_name']
            if page_2_store_data['selected_file_name']:
                    name_value = page_2_store_data['selected_file_name']
                    # Leer el archivo Excel
                    file_path = os.path.join(f'configurations/{selected_file_name}.xlsx')
                    df = pd.read_excel(file_path)

                    for i, col in enumerate(ids):
                        if col in df.columns:
                            inputs_values[i] = df[col].iloc[0]

                    styles = [{'border': '2px solid green'}] * len(ids)
                
                    return styles + ['', '', {}] + [name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]
        except: 
            selected_file_name = None
            
       
    if triggered_id == 'open-modal-button':
        file_names = list_files()
        options = [{'label': name, 'value': name} for name in file_names]
        return styles + ['', '', {'display': 'block'}, name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [options]

    if triggered_id == 'cancel-button':
        return styles + ['', '', {'display': 'none'}, name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

    if triggered_id == 'save-button':
        all_filled = True

        if save_clicks is None:
            return styles + ['', '', {}] + [name_value] + inputs_values + [[]]

        for i, value in enumerate(values[:-2]):  # Excluir los últimos dos valores que son 'name-input' y 'file-dropdown'
            if value is None or value == '':
                styles[i] = {'border': '2px solid red'}
                all_filled = False
            else:
                styles[i] = {'border': '2px solid green'}

        if not all_filled:
            return styles + ['Please fill all the fields.', '', {}] + [name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

        if values[-2] is None or values[-2] == '':
            return styles + ['Please introduce a valid configuration name', '', {'color': 'red'}] + [name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

        file_names = list_files()
        if values[-2] in file_names:
            return styles + ['Please introduce a non-existent configuration name', '', {'color': 'red'}] + [name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

        # Crear el DataFrame
        parameters = pd.DataFrame([{
            'Tc': values[0],
            'Masa_molar': values[1]/1000,
            'gamma': values[2],
            'rho_pr': values[3],
            'Rend': values[4]/100,
            'a': values[5],
            'n': values[6],
            'Mcp': values[7],
            'Rcg_cp': values[8]/1000,
            'Cd': values[9],
            'Re': values[10] * 1000,
            'g0': values[11],
            'Ra': values[12],
            'Mc': values[13],
            't0': values[14],
            'tf': values[15],
            'N': values[16]
        }])

        if not os.path.exists('configurations'):
            os.makedirs('configurations')

        parameters.to_excel(f'configurations/{values[-2]}.xlsx', index=False)

        return styles + ['', 'Parameters saved successfully!', {'color': 'green'}]+[name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

    if triggered_id == 'accept-button' and file_dropdown_value:
        try:
            # Leer el archivo Excel
            file_path = os.path.join(f'configurations/{file_dropdown_value}.xlsx')
            df = pd.read_excel(file_path)

            for i, col in enumerate(ids):
                if col in df.columns:
                    inputs_values[i] = df[col].iloc[0]

            name_value = file_dropdown_value
            selected_file_name = file_dropdown_value  # Guardar el nombre del archivo seleccionado en la variable global
            styles = [{'border': '2px solid green'}] * len(ids)
            success_message = f'Loaded file successfully: {file_dropdown_value}'
            return styles + [error_message, success_message, {'display': 'none'}, name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]
        except Exception as e:
            return styles + ['There was an error processing this file.', '', {'display': 'none'}, name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

    return styles + ['', '', {}] + [name_value, {'inputs_values': inputs_values, 'name_value': name_value, 'selected_file_name': name_value}] + inputs_values + [[]]

# Callback para actualizar el label en la página 1
@app.callback(
    Output('selected-file-label', 'children'),
    Input('url', 'pathname'),
    State('page-2-store', 'data')
)
def update_selected_file_label(pathname, page_2_store_data):
    if pathname == '/simulations':
        try:
            output = page_2_store_data['selected_file_name']
            
        except: 
            output = ''
            
        return output
        
    return ''

# Callback para actualizar el gráfico
@app.callback(
    [Output('contour-graph', 'figure'),
     Output('download-nozzle-button', 'disabled'),
     Output('download-nozzle-button', 'style'),
     Output('calculate-button', 'disabled'),
     Output('calculate-button', 'style'),
     ],
    
    [Input('rg-input', 'value'),
     Input('rs-input', 'value'),
     Input('R-input', 'value'),
     Input('R0-input', 'value'),
     Input('L-input', 'value'),
     Input('N_sc-input', 'value'),
     Input('nozzle-type-dropdown', 'value'),
     Input('3d-toggle', 'value'),
     Input('show-axis-toggle', 'value')],
    
    [State('page-2-store', 'data')]
    
)

def update_graph(rg, rs, R, R0, L, N_sc, nozzle_type, is_3d, is_axis, page_2_store_data):
    if 'axis' in is_axis:
        axis=True
    else:
        axis=False
    
    try:
        config_filename = page_2_store_data['selected_file_name']
    except:
        config_filename = None
    if None in [rg, rs, R, R0, L, N_sc, config_filename]:
        
            
        calculate_disabled = True
        calculate_style = { 'backgroundColor': 'lightgray', 'border':'black'}
    else:
        calculate_disabled = False
        calculate_style = {}
    if None in [rg, rs, R, R0, L, nozzle_type]:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=[], y=[], mode='lines'))
        fig.update_layout(
            xaxis_title='Radius (m)',
            yaxis_title='Longitud axial (m)',
            template='plotly_white',
            xaxis=dict(
                scaleanchor='y',
                scaleratio=1,
                constrain='domain',
                showgrid=False,  # Ocultar líneas cuadriculadas
                zeroline=False,  # Ocultar línea del eje
                visible=False    # Ocultar el eje
            ),
            yaxis=dict(
                scaleanchor='x',
                scaleratio=1,
                constrain='domain',
                showgrid=False,  # Ocultar líneas cuadriculadas
                zeroline=False,  # Ocultar línea del eje
                visible=False    # Ocultar el eje
            ),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)'
        )

        return fig, True, { 'backgroundColor': 'lightgray', 'border':'black'}, calculate_disabled, calculate_style 

    k = 1.21  # Ratio de calores específicos, típico
    aratio = (rs / rg) ** 2
    l_percent = 80  # Porcentaje de longitud de la tobera

    angles, contour = bell_contour.bell_nozzle(k, aratio, rg, l_percent)

    ye, xe, nye, ye2, xe2, nye2, ybell, xbell, nybell = contour

    ye = np.array([-y for y in ye])
    ye2 = np.array([-y for y in ye2])
    ybell = np.array([-y for y in ybell])

    ye = ye + np.abs(np.min(ybell))
    ye2 = ye2 + np.abs(np.min(ybell))
    ybell = ybell + np.abs(np.min(ybell))

    point1 = (xe[0], ye[0])
    point2 = (R, max(ye))
    point3 = (R, max(ye) + L)
    point4 = (0, max(ye) + L)

    segment1 = {'x': [point1[0], point2[0]], 'y': [point1[1], point2[1]]}
    segment2 = {'x': [point2[0], point3[0]], 'y': [point2[1], point3[1]]}
    segment3 = {'x': [point3[0], point4[0]], 'y': [point3[1], point4[1]]}

    if '3d' in is_3d:
        # Crear una superficie de revolución
        theta = np.linspace(0, 2 * np.pi, 100)
        
        def revolve(x, y):
            theta_grid, r_grid = np.meshgrid(theta, x)
            x_grid = r_grid * np.cos(theta_grid)
            y_grid = r_grid * np.sin(theta_grid)
            z_grid = np.tile(y, (len(theta), 1)).T
            return x_grid, y_grid, z_grid

        # Revolución de las curvas
        xe_grid, ye_grid, ze_grid = revolve(xe, ye)
        xe2_grid, ye2_grid, ze2_grid = revolve(xe2, ye2)
        xbell_grid, ybell_grid, zbell_grid = revolve(xbell, ybell)

        # Revolución de los segmentos
        segment_grids = []
        segments=[segment1, segment2, segment3]
        for segment in segments:
            sx, sy = segment['x'], segment['y']
            sx_grid, sy_grid, sz_grid = revolve(sx, sy)
            segment_grids.append((sx_grid, sy_grid, sz_grid))

        fig = go.Figure(data=[
            go.Surface(x=xe_grid, y=ze_grid, z=ye_grid, showscale=False, name='Entrante de la garganta', hoverinfo='none',
        contours=dict(
            x=dict(highlight=False),
            y=dict(highlight=False),
            z=dict(highlight=False)
        )),
            go.Surface(x=xe2_grid, y=ze2_grid, z=ye2_grid, showscale=False, name='Salida de la garganta', hoverinfo='none',
        contours=dict(
            x=dict(highlight=False),
            y=dict(highlight=False),
            z=dict(highlight=False)
        )),
            go.Surface(x=xbell_grid, y=zbell_grid, z=ybell_grid, showscale=False, name='Contorno de la campana', hoverinfo='none',
        contours=dict(
            x=dict(highlight=False),
            y=dict(highlight=False),
            z=dict(highlight=False)
        ))
        ])

        # Añadir superficies de los segmentos
        for idx, (sx_grid, sy_grid, sz_grid) in enumerate(segment_grids):
            fig.add_trace(go.Surface(x=sx_grid, y=sz_grid, z=sy_grid, showscale=False, name=f'Segmento {idx + 1}', hoverinfo='none', contours=dict(
            x=dict(highlight=False),
            y=dict(highlight=False),
            z=dict(highlight=False), 
        )))

        fig.update_layout(
            scene=dict(
                        xaxis=dict(
                            showgrid=False,  # Ocultar líneas cuadriculadas
                            zeroline=False,  # Ocultar línea del eje
                            visible=axis    # Ocultar el eje
                        ),
                        yaxis=dict(
                            showgrid=False,  # Ocultar líneas cuadriculadas
                            zeroline=False,  # Ocultar línea del eje
                            visible=axis    # Ocultar el eje
                        ),
                        zaxis=dict(
                            showgrid=False,  # Ocultar líneas cuadriculadas
                            zeroline=False,  # Ocultar línea del eje
                            visible=axis    # Ocultar el eje
                        ),
                        ),
            plot_bgcolor='rgba(0,0,0,0)',  # Fondo del área de trazado transparente
            paper_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig, False, {'backgroundColor': ''}, calculate_disabled, calculate_style
    else:
        fig = go.Figure()

        # Función para reflejar los valores de x respecto al eje y
        def reflect_x(x_values):
            return [-x for x in x_values]

        # Entrante de la garganta
        fig.add_trace(go.Scatter(x=xe, y=ye, mode='lines', name='Entrante de la garganta', line=dict(color='black'), showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(xe), y=ye, mode='lines', line=dict(color='black'), showlegend=False))

        # Salida de la garganta
        fig.add_trace(go.Scatter(x=xe2, y=ye2, mode='lines', name='Salida de la garganta', line=dict(color='black'), showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(xe2), y=ye2, mode='lines', line=dict(color='black'), showlegend=False),)

        # Contorno de la campana
        fig.add_trace(go.Scatter(x=xbell, y=ybell, mode='lines', name='Contorno de la campana', line=dict(color='black'), showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(xbell), y=ybell, mode='lines', line=dict(color='black'), showlegend=False))

        # Segmentos adicionales
        fig.add_trace(go.Scatter(x=segment1['x'], y=segment1['y'], mode='lines', name='Segment 1', line=dict(color='black'), showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(segment1['x']), y=segment1['y'], mode='lines', line=dict(color='black'),showlegend=False))

        fig.add_trace(go.Scatter(x=segment2['x'], y=segment2['y'], mode='lines', name='Segment 2', line=dict(color='black'), showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(segment2['x']), y=segment2['y'], mode='lines', line=dict(color='black'), showlegend=False))

        fig.add_trace(go.Scatter(x=segment3['x'], y=segment3['y'], mode='lines', name='Segment 3', line=dict(color='black'),showlegend=False))
        fig.add_trace(go.Scatter(x=reflect_x(segment3['x']), y=segment3['y'], mode='lines', line=dict(color='black'), showlegend=False))

        # Añadir el rectángulo sin bordes
        fig.add_shape(
            type="rect",
            x0=-R, y0=max(ye)+0.1, x1=-R0-0.1, y1=max(ye)+L-0.1,
            line=dict(color="RoyalBlue"),  # Esto quitará el borde del rectángulo
            fillcolor="LightSkyBlue",
        )
        fig.add_shape(
            type="rect",
            x0=R0, y0=max(ye)+0.1, x1=R-0.1, y1=max(ye)+L-0.1,
            line=dict(color="RoyalBlue"),
            fillcolor="LightSkyBlue",
        )

        # Encontrar el rango máximo necesario
        max_x = max(max(xe), max(xe2), max(xbell), R)
        max_y = 1.1*max(max(ye), max(ye2), max(ybell), point4[1])

        # Ajustar el rango de los ejes para que sean iguales y se vean todos los elementos
        max_range = max(max_x, max_y)

        fig.update_layout(
            xaxis_title='Radius (mm)',
            yaxis_title='Length (mm)',
            template='plotly_white',
            xaxis=dict(
                range=[-max_range, max_range],
                scaleanchor='y',
                scaleratio=1,
                constrain='domain',
                showgrid=False,  # Ocultar líneas cuadriculadas
                zeroline=False,  # Ocultar línea del eje
                visible=axis    # Ocultar el eje
            ),
            yaxis=dict(
                range=[0, max_range],
                scaleanchor='x',
                scaleratio=1,
                constrain='domain',
                showgrid=False,  # Ocultar líneas cuadriculadas
                zeroline=False,  # Ocultar línea del eje
                visible=axis    # Ocultar el eje
            ),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)'
        )

    return fig, False, {'backgroundColor': ''}, calculate_disabled, calculate_style

@app.callback(
    Output('download-nozzle-contour', 'data'),
    Input('download-nozzle-button', 'n_clicks'),
    State('rg-input', 'value'),
    State('rs-input', 'value'),
    prevent_initial_call=True
)
def download_contour_nozzle(n_clicks, rg, rs):
    if rg is None or rs is None:
        return dash.no_update

    k = 1.21  # Ratio de calores específicos, típico
    aratio = (rs / rg) ** 2
    l_percent = 80  # Porcentaje de longitud de la tobera

    # Supongo que tienes una función llamada bell_nozzle que genera el contorno
    # Sustituye esto por la implementación correcta
    angles, contour = bell_contour.bell_nozzle(k, aratio, rg, l_percent)

    ye, xe, nye, ye2, xe2, nye2, ybell, xbell, nybell = contour

    x = np.concatenate((xe, xe2, xbell))
    y = np.concatenate((ye, ye2, ybell))

    contour_df = pd.DataFrame({'x': x, 'y': y})

    # Usamos un buffer de memoria para evitar el uso de archivos temporales
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        contour_df.to_excel(writer, index=False, sheet_name='Nozzle Contour')

    buffer.seek(0)
    
    return dcc.send_bytes(buffer.getvalue(), "nozzle_contour.xlsx")
    
# Callback to update graphs
@app.callback(
    [Output(f'graph{i}', 'figure') for i in range(1, 13)]+[Output('graphs-container', 'style'), Output('open-modal2-button', 'style')],
    [Input('calculate-button', 'n_clicks')],
    [State('rg-input', 'value'),
     State('rs-input', 'value'),
     State('R-input', 'value'),
     State('R0-input', 'value'),
     State('L-input', 'value'),
     State('N_sc-input', 'value'),
     State('page-2-store', 'data')]
)

def update_output(n_clicks, rg, rs, R, R0, L, N_sc, page_2_store_data):
    try:
        config_filename = page_2_store_data['selected_file_name']
    except:
        config_filename = None
    
    if n_clicks is None or config_filename is None:
        return [go.Figure() for _ in range(12)]+[{'display':'none'}]+[{'display':'none'}]
    else:
        config_path = os.path.join(f'configurations/{config_filename}.xlsx')
        parameters = pd.read_excel(config_path)
        
        if N_sc>parameters['N'].iloc[0]:
            N_sc = parameters['N'].iloc[0]
        
        
        rocket_calculator.rocket(R/1000, R0/1000, L/1000, rg/1000, rs/1000, N_sc, parameters)
        file_path = os.path.join(os.getcwd(), "temp", "data.xlsx")
        
        parameters_sim = pd.DataFrame([{
            'rg': rg,
            'rs': rs,
            'R': R,
            'R0': R0,
            'L': L,
            'N_sc': N_sc,
        }])
        
        parameters_full = pd.concat([parameters, parameters_sim], axis=1)
        
        
        try:
            # Cargar el archivo Excel existente
            book = load_workbook(file_path)
            
            if 'parameters' in book.sheetnames:
                del book['parameters']
                
            book.save(file_path)

            # Ahora, abrir el ExcelWriter para agregar la nueva hoja
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                pd.DataFrame(parameters_full).to_excel(writer, sheet_name="parameters", index=False)
    
            engine_scatter_data = pd.read_excel(file_path, sheet_name='engine_scatter')
            attitude_scatter_data = pd.read_excel(file_path, sheet_name='attitude_scatter')
            
        except Exception as e:
            return [go.Figure(data=[go.Scatter(x=[], y=[], mode='markers', text=str(e))]) for _ in range(12)]+[{'display':'block'}]+[{'display':'block'}]
        
        
        line_colors = ['blue', 'green', 'red', 'cyan', 'magenta', 'brown', 'black', 'purple', 'orange', 'blue', 'pink', 'red']
        point_colors = ['rgb({},{},{})'.format(r, g, b) for r, g, b in zip(engine_scatter_data['color_s_1']*255, engine_scatter_data['color_s_2']*255, engine_scatter_data['color_s_3']*255)]
        graphs = ['Pc_s', 'E_s', 'Ts_s', 'r_s', 'Ms_s', 'Vs_s', 'Rp_s', 'Rp_s', 'Fa_s', 'h_s', 'der_s', 'der2_s']
        graphs_titles = ['Chamber pressure', 'Thrust', 'Exit temperature', 'Fuel radius', 'Exit Mach', 'Exit velocity','Recesion rate', 'Mass flux', 'Drag_force', 'Altitude', 'Rocket velocity', 'Rocket acceleration']
        y_axis = ['Pc (atm)',  'E (kN)', 'Ts (k)', 'r (mm)', 'Mach', 'Vs (m/s)', 'Rp (mm/s)','Mass flux (kg/s)', 'Fa (N)', 'h(m)', 'v(m/s)', 'a(m/s^2)' ]
        unit_factor = [1/101325, 1/1000, 1, 1000, 1, 1, 1000, 1, 1, 1, 1, 1]
        figures = []
        for i in range(7):
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=engine_scatter_data['t_s'], y=engine_scatter_data[graphs[i]]*unit_factor[i], mode='lines+markers', name=graphs_titles[i], line=dict(color='black'), marker=dict(color=point_colors)))
            fig.update_layout(title=dict(text=graphs_titles[i], x=0.5, font=dict(size=20, color='black', family='Arial')), xaxis=dict(title='Time (s)', showgrid=True, gridcolor='lightgrey', linecolor='black'), yaxis=dict(title=y_axis[i], showgrid=True, gridcolor='lightgrey', linecolor='black'), plot_bgcolor='rgb(240, 248, 255)',  paper_bgcolor='rgb(240, 248, 255)' )
            figures.append(fig)
        
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=engine_scatter_data['t_s'], y=engine_scatter_data['Ge_s']*unit_factor[i+1], mode='lines', name='Created mass flow', line=dict(color='orange')))
        fig.add_trace(go.Scatter(x=engine_scatter_data['t_s'], y=engine_scatter_data['Gs_s']*unit_factor[i+1], mode='lines', name='Exiting mass flow', line=dict(color='blue')))
        fig.update_layout(title=dict(text=graphs_titles[i+1], x=0.5, font=dict(size=20, color='black', family='Arial')), xaxis=dict(title='Time (s)', showgrid=True, gridcolor='lightgrey', linecolor='black'), yaxis=dict(title=y_axis[i+1], showgrid=True, gridcolor='lightgrey', linecolor='black'), plot_bgcolor='rgb(240, 248, 255)',  paper_bgcolor='rgb(240, 248, 255)' )
        figures.append(fig)
        
        point_colors2= ['rgb({},{},{})'.format(r, g, b) for r, g, b in zip(attitude_scatter_data['color_s2_1']*255, attitude_scatter_data['color_s2_2']*255, attitude_scatter_data['color_s2_3']*255)]
        for i in range(4):
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=attitude_scatter_data['t_s2'], y=attitude_scatter_data[graphs[8+i]]*unit_factor[8+i], mode='lines+markers', name=graphs_titles[8+i], line=dict(color='black'), marker=dict(color=point_colors2)))
            fig.update_layout(title=dict(text=graphs_titles[8+i], x=0.5, font=dict(size=20, color='black', family='Arial')), xaxis=dict(title='Time (s)', showgrid=True, gridcolor='lightgrey', linecolor='black'), yaxis=dict(title=y_axis[8+i], showgrid=True, gridcolor='lightgrey', linecolor='black'), plot_bgcolor='rgb(240, 248, 255)',  paper_bgcolor='rgb(240, 248, 255)' )
            figures.append(fig)
        
        return figures+[{'display':'block'}]+[{'display':'block'}]

@app.callback(
    [
        Output('modal2', 'style'),
        Output('save_filename-input', 'style'),
        Output('name-exists-warning', 'children'),
        Output('success-saving-message', 'children')
    ],
    [
        Input('open-modal2-button', 'n_clicks'),
        Input('accept-button-modal2', 'n_clicks'),
        Input('cancel-button-modal2', 'n_clicks')
    ],
    [State('save_filename-input', 'value')]
)
def update_output(open_clicks, accept_clicks, cancel_clicks, input_value):
    ctx = dash.callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]

    existing_names = ['name1', 'name2', 'name3']  # Lista de nombres existentes

    # Abrir el modal
    if triggered_id == 'open-modal2-button':
        return {'display': 'block'}, {'border': '1px solid black'}, '', ''

    # Cerrar el modal
    if triggered_id == 'cancel-button-modal2':
        return {'display': 'none'}, {'border': '1px solid black'}, '', ''

    # Validar nombre en el campo de entrada
    if triggered_id == 'accept-button-modal2':
        if input_value in existing_names:
            return {'display': 'block'}, {'border': '2px solid red'}, 'Name already exists. Please select a new one.', ''
        else:
            if not os.path.exists('results'):
                os.makedirs('results')
                
            origen = os.path.join(os.getcwd(), 'temp', 'data.xlsx')
            destino = os.path.join(os.getcwd(), 'results', f'{input_value}.xlsx')
            shutil.move(origen, destino)
            return {'display': 'none'}, {'border': '1px solid black'}, '', f'File saved succesfully in {input_value}'

    return {'display': 'none'}, {'border': '1px solid black'}, ''
    
if __name__ == '__main__':
    app.run_server(debug=True)
